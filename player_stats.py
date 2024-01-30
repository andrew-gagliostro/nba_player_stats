from datetime import datetime, timedelta, time
import json
from nba_api.stats.endpoints import commonplayerinfo, leagueseasonmatchups, leaguegamefinder, teamgamelogs, playergamelogs, teamdashlineups, leaguedashplayerbiostats
from nba_api.stats.static import players, teams
from collections import Counter, defaultdict
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment

def get_player_ids(player_names):
    target_players = [name.strip() for name in player_names.split(",")]
    player_ids = dict()
    players = leaguedashplayerbiostats.LeagueDashPlayerBioStats().get_normalized_dict()["LeagueDashPlayerBioStats"]
    # with open('all_players.json', 'w') as f:
    #     json.dump(players, f)

    for player in players:
        if player['PLAYER_NAME'] in target_players:
            player_ids[player['PLAYER_NAME']] = player['PLAYER_ID']

    # print(player_ids)
    return player_ids

def check_games(players_ids, player_data, games):
    games_counter = Counter()
    for player_id in players_ids:
        for player in player_data["PlayerGameLogs"]:
            if player["PLAYER_ID"] == player_id:
                games_counter[player["GAME_ID"]] += 1
    
    return [game for game, count in games_counter.items() if count == len(players_ids)]

while True:
    player_names = input("Enter player names separated by commas in the format 'P1, P2, P3': ")
    if player_names == "":
        break
    season = input("Enter the season in the format 'YYYY-YY': ")

    verbose_output = input("Do you want verbose output for player activity of each game (y/n)?: ")


    # Assuming games_data and player_data are your datasets
    games_data = teamgamelogs.TeamGameLogs(season_nullable=season).get_normalized_json()
    # players_data = player_info = playergamelogs.PlayerGameLogs(player_id_nullable=201142, season_nullable="2023-24").get_normalized_json()

    player_ids = get_player_ids(player_names)
    all_player_info = playergamelogs.PlayerGameLogs(season_nullable=season).get_normalized_dict()

    games_counter = Counter()
    games_dict = dict()
    game_player_activity_dict = defaultdict(dict)
    lineups = defaultdict(str)
    stats_types = ['MIN', 'FGM', 'FGA', 'FG_PCT', 'FG3M', 'FG3A', 'FG3_PCT', 'FTM', 'FTA', 'FT_PCT','OREB', 'DREB', 'REB', 'AST', 'TOV', 'STL', 'BLK', 'BLKA', 'PF', 'PFD','PTS']
    for player_name, player_id in player_ids.items():
        games = playergamelogs.PlayerGameLogs(player_id_nullable=player_id, season_nullable=season).get_normalized_dict()["PlayerGameLogs"]
        for game in games:
            if game["MIN"] != 0:
                game_id = game["GAME_ID"]
                this_player_activity = ""
                for stat in stats_types:
                    if game[stat] is not None:
                        this_player_activity += f"{stat}: {round(game[stat], 2)}\n"
                game_player_activity_dict[game_id][player_name] = this_player_activity
                games_counter[game_id] += 1
                if games_counter[game_id] == len(player_ids):
                    games_dict[game_id] = game
                    lineups[game_id] = teamdashlineups.TeamDashLineups(team_id=game["TEAM_ID"], game_id_nullable=game_id).get_normalized_dict()["Lineups"][0]["GROUP_NAME"]


    with open('games_data.json', 'w') as f:
        json.dump(games_data, f)

    breakdown = []
    wins = 0
    losses = 0

    for game_id, game in games_dict.items():
        date = datetime.strptime(game["GAME_DATE"],"%Y-%m-%dT%H:%M:%S").strftime("%m-%d-%Y")
        if game["WL"] == "W":
            wins += 1
        else:
            losses += 1
        summary = [date, game["MATCHUP"], game["WL"], lineups[game_id]]
        if verbose_output.lower() == 'y':
            for player_name in player_ids.keys():
                player_game_data = game_player_activity_dict[game_id][player_name]
                summary.append(player_game_data)
        breakdown.append(summary)

    def get_date(game):
        date_str = game[0]
        return datetime.strptime(date_str, '%m-%d-%Y')

    sorted_games = sorted(breakdown, key=get_date)
    filename_excel = "".join(player_names.split(", ")).replace(" ", "") + datetime.now().strftime("%m-%d-%Y") + ".xlsx"
    df_columns = ["Date", "Matchup", "Win/Loss", "Starters"]


    if verbose_output.lower() == 'y':
        additional_columns = [f"{name} Game Stats" for name in player_ids.keys()]
        df_columns = df_columns + additional_columns

    if len(df_columns) != len(sorted_games[0]):
        print(f"Column number does not match: {len(df_columns)} vs {len(sorted_games[0])}")
        continue

    df_games = pd.DataFrame(sorted_games, columns=df_columns)

    summary_df = pd.DataFrame([["Players", "Season", "Games", "Record"],
                               [player_names, season, f"{wins + losses}", f"{wins}-{losses}"]],
                            columns=None)
    with pd.ExcelWriter(filename_excel, engine='openpyxl') as writer:
        summary_df.to_excel(writer, index=False, header=False, sheet_name='Sheet1')
        df_games.to_excel(writer, index=False, sheet_name='Sheet1', startrow=4)

    from openpyxl import load_workbook
    wb = load_workbook(filename_excel)
    sheet = wb['Sheet1']

    for row in sheet.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal='center')

    for row in sheet.iter_rows(min_row=6):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)

    bold_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = bold_font

    fill = PatternFill(start_color="FFC7CE",
                    end_color="FFC7CE",
                    fill_type="solid")

    for cell in sheet.iter_cols(min_col=1, max_col=len(df_columns), min_row=5, max_row=5):
        cell[0].fill = fill

    for cell in sheet.iter_cols(min_col=1, max_col=4, min_row=1, max_row=1):
        cell[0].fill = fill

    i = 0
    for i, column_cells in enumerate(sheet.columns):
        max_length = 0
        column = get_column_letter(column_cells[0].column)  # convert column index to letter

        # Set width to 50 for columns past the third one
        if i >= 3:
            sheet.column_dimensions[column].width = 30
            continue

        for cell in column_cells:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    
    # if verbose_output.lower() == 'y':
    #     stats_col_idxs = range(4, 3+len(player_ids))
    #     for row in sheet.iter_rows(min_row=6, max_row=sheet.max_row, min_col=4, max_col=3+len(player_ids)):
    #         for cell in row:
    #             cell.alignment = Alignment(wrap_text=True)

    wb.save(filename_excel)
    
    print(f"Output saved as {filename_excel}")
    print("Press enter To exit, or enter player names to continue with another query")


